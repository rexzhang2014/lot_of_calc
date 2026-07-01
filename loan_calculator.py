#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
贷款计算器（含多次提前还款功能）
================================
支持等额本息和等额本金两种还款方式。
支持多次提前还款，每次可独立选择策略：
  1. 缩短期限（月供不变，期数变少）
  2. 降低月供（期数不变，月供变少）

核心 API:
  - calculate_loan(principal, annual_rate, total_months, method)
  - calculate_multiple_prepayments(original_result, prepayments)
  - export_to_csv(result, filepath, multi_prepay=None)
  - export_to_xlsx(result, filepath, multi_prepay=None)

用法:
    python loan_calculator.py
    # 按提示输入参数即可

模块导入示例:
    from loan_calculator import (
        calculate_loan,
        calculate_multiple_prepayments,
        export_to_csv,
        export_to_xlsx,
        print_summary,
        print_schedule,
        print_multiple_prepayment_summary,
    )

    # 1. 计算原始贷款
    result = calculate_loan(
        principal=650000,
        annual_rate=0.0391,
        total_months=324,
        method='等额本息',
    )

    # 2. 定义多次提前还款计划（用户输入的参数）
    prepayments = [
        {"period": 60,  "amount": 100000, "strategy": "缩短期限"},
        {"period": 100, "amount": 100000, "strategy": "降低月供"},
    ]

    # 3. 计算多次提前还款后的结果
    multi = calculate_multiple_prepayments(result, prepayments)

    # 4. 导出
    export_to_xlsx(result, "还款计划.xlsx", multi)
"""

import csv
import math
import os
import sys
from copy import deepcopy
from dataclasses import dataclass, field
from typing import Dict, List, Literal, Optional, Tuple


# =============================================================================
# 数据模型
# =============================================================================

@dataclass
class PaymentRecord:
    """单期还款记录"""
    period: int              # 期数
    monthly_payment: float   # 月供金额
    principal_paid: float    # 本月所还本金
    interest_paid: float     # 本月所还利息
    remaining_principal: float  # 上月剩余本金（还款前）
    remaining_interest: float   # 上月剩余利息（还款前）
    is_prepayment: bool = False  # 是否为提前还款当期
    prepay_amount: float = 0.0   # 当期提前还款金额（普通月供为0）
    strategy: str = ""           # 提前还款策略


@dataclass
class LoanResult:
    """贷款计算结果"""
    principal: float         # 贷款本金
    annual_rate: float       # 合同年利率
    total_months: int        # 总期数（月）
    method: str              # 还款方式
    monthly_rate: float      # 月利率
    total_payment: float     # 总还款额
    total_interest: float    # 总利息
    records: List[PaymentRecord]  # 还款计划明细


@dataclass
class PrepaymentEvent:
    """单次提前还款事件"""
    prepay_period: int       # 在第几期后提前还款
    prepay_amount: float     # 提前还款金额
    strategy: str            # 还款策略


@dataclass
class MultiplePrepaymentResult:
    """多次提前还款计算结果"""
    original_result: LoanResult           # 原始贷款结果
    prepayments: List[PrepaymentEvent]  # 提前还款事件列表

    total_prepay_amount: float   # 累计提前还款总额
    final_months: int            # 最终总期数
    final_total_payment: float   # 最终总还款额（含所有提前还款）
    final_total_interest: float  # 最终总利息
    interest_saved: float        # 相比不提前还款节省的利息

    prepay_details: List[Dict] = field(default_factory=list)
    full_records: List[PaymentRecord] = field(default_factory=list)


# =============================================================================
# 核心计算函数（私有）
# =============================================================================

def _calc_equal_principal_and_interest(
    principal: float,
    annual_rate: float,
    monthly_rate: float,
    total_months: int,
    start_period: int = 1,
    is_prepayment: bool = False,
    prepay_amount: float = 0.0,
    strategy: str = "",
) -> List[PaymentRecord]:
    """等额本息还款方式计算（内部通用）"""
    if monthly_rate == 0:
        monthly_payment = principal / total_months
    else:
        monthly_payment = (
            principal
            * monthly_rate
            * (1 + monthly_rate) ** total_months
            / ((1 + monthly_rate) ** total_months - 1)
        )

    records = []
    remaining_principal = principal

    for i in range(total_months):
        period = start_period + i
        prev_remaining = remaining_principal
        interest_paid = remaining_principal * monthly_rate
        principal_paid = monthly_payment - interest_paid
        remaining_principal -= principal_paid

        # 剩余利息（理论值）
        remaining_months = total_months - i
        if monthly_rate == 0:
            remaining_interest = 0.0
        elif remaining_principal > 0 and remaining_months > 0:
            future_monthly = (
                remaining_principal
                * monthly_rate
                * (1 + monthly_rate) ** remaining_months
                / ((1 + monthly_rate) ** remaining_months - 1)
            )
            remaining_interest = future_monthly * remaining_months - remaining_principal
        else:
            remaining_interest = 0.0

        records.append(
            PaymentRecord(
                period=period,
                monthly_payment=round(monthly_payment, 2),
                principal_paid=round(principal_paid, 2),
                interest_paid=round(interest_paid, 2),
                remaining_principal=round(prev_remaining, 2),
                remaining_interest=round(remaining_interest, 2),
                is_prepayment=is_prepayment and i == 0,
                prepay_amount=prepay_amount if (is_prepayment and i == 0) else 0.0,
                strategy=strategy if (is_prepayment and i == 0) else "",
            )
        )

    return records


def _calc_equal_principal(
    principal: float,
    annual_rate: float,
    monthly_rate: float,
    total_months: int,
    start_period: int = 1,
    is_prepayment: bool = False,
    prepay_amount: float = 0.0,
    strategy: str = "",
) -> List[PaymentRecord]:
    """等额本金还款方式计算（内部通用）"""
    monthly_principal = principal / total_months
    records = []
    remaining_principal = principal

    for i in range(total_months):
        period = start_period + i
        prev_remaining = remaining_principal
        interest_paid = remaining_principal * monthly_rate
        principal_paid = monthly_principal
        monthly_payment = principal_paid + interest_paid
        remaining_principal -= principal_paid

        # 剩余利息
        remaining_months = total_months - i
        if monthly_rate == 0:
            remaining_interest = 0.0
        elif remaining_principal > 0 and remaining_months > 0:
            remaining_interest = sum(
                (remaining_principal - monthly_principal * j) * monthly_rate
                for j in range(remaining_months)
            )
        else:
            remaining_interest = 0.0

        records.append(
            PaymentRecord(
                period=period,
                monthly_payment=round(monthly_payment, 2),
                principal_paid=round(principal_paid, 2),
                interest_paid=round(interest_paid, 2),
                remaining_principal=round(prev_remaining, 2),
                remaining_interest=round(remaining_interest, 2),
                is_prepayment=is_prepayment and i == 0,
                prepay_amount=prepay_amount if (is_prepayment and i == 0) else 0.0,
                strategy=strategy if (is_prepayment and i == 0) else "",
            )
        )

    return records


# =============================================================================
# 公共 API
# =============================================================================

def calculate_loan(
    principal: float,
    annual_rate: float,
    total_months: int,
    method: Literal["等额本息", "等额本金"],
) -> LoanResult:
    """
    计算原始贷款还款计划。

    参数:
        principal:    贷款本金
        annual_rate:  合同年利率（如 0.043 表示 4.3%）
        total_months: 总期数（月）
        method:       还款方式，'等额本息' 或 '等额本金'

    返回:
        LoanResult 对象
    """
    monthly_rate = annual_rate / 12

    if method == "等额本息":
        records = _calc_equal_principal_and_interest(
            principal, annual_rate, monthly_rate, total_months
        )
    elif method == "等额本金":
        records = _calc_equal_principal(
            principal, annual_rate, monthly_rate, total_months
        )
    else:
        raise ValueError("还款方式必须是 '等额本息' 或 '等额本金'")

    total_payment = sum(r.monthly_payment for r in records)
    total_interest = sum(r.interest_paid for r in records)

    return LoanResult(
        principal=principal,
        annual_rate=annual_rate,
        total_months=total_months,
        method=method,
        monthly_rate=monthly_rate,
        total_payment=round(total_payment, 2),
        total_interest=round(total_interest, 2),
        records=records,
    )


def calculate_multiple_prepayments(
    original_result: LoanResult,
    prepayments: List[Dict],
) -> MultiplePrepaymentResult:
    """
    多次提前还款计算。

    参数:
        original_result: 原始贷款计算结果（由 calculate_loan 生成）
        prepayments: 提前还款事件列表，每项为字典:
            {
                "period":   int,    # 在第几期后提前还款（必须递增）
                "amount":   float,  # 提前还款金额
                "strategy": str,    # '缩短期限' 或 '降低月供'
            }

    返回:
        MultiplePrepaymentResult 对象
    """
    # 排序并验证
    sorted_prepays = sorted(prepayments, key=lambda x: x["period"])
    for i, p in enumerate(sorted_prepays):
        if p["period"] < 1 or p["period"] >= original_result.total_months:
            raise ValueError(
                f"第 {i + 1} 次提前还款期数必须在 1 到 {original_result.total_months - 1} 之间"
            )
        if p["amount"] <= 0:
            raise ValueError(f"第 {i + 1} 次提前还款金额必须大于 0")

    method = original_result.method
    monthly_rate = original_result.monthly_rate
    annual_rate = original_result.annual_rate

    full_records: List[PaymentRecord] = []
    prepay_details: List[Dict] = []
    total_prepay_amount = 0.0

    # 当前状态
    current_period = 1
    current_principal = original_result.principal
    current_records = original_result.records
    current_total_months = original_result.total_months

    for idx, prepay in enumerate(sorted_prepays):
        prepay_period = prepay["period"]
        prepay_amount = prepay["amount"]
        strategy = prepay["strategy"]

        # ---- 提前还款前的正常还款段 ----
        start_idx = current_period - 1
        end_idx = prepay_period

        if start_idx < len(current_records):
            segment = []
            for i in range(start_idx, min(end_idx, len(current_records))):
                r = deepcopy(current_records[i])
                r.period = i - start_idx + current_period
                segment.append(r)

            full_records.extend(segment)

            if segment:
                last = segment[-1]
                current_principal = last.remaining_principal - last.principal_paid
            else:
                current_principal = original_result.principal

        # ---- 执行提前还款 ----
        if prepay_amount >= current_principal:
            raise ValueError(
                f"第 {idx + 1} 次提前还款金额 {prepay_amount:,.2f} "
                f"超过剩余本金 {current_principal:,.2f}"
            )

        remaining_before = current_principal
        total_prepay_amount += prepay_amount
        current_principal -= prepay_amount

        # ---- 计算新的还款期数 ----
        remaining_original_months = current_total_months - prepay_period

        if strategy == "缩短期限":
            if method == "等额本息":
                if monthly_rate == 0:
                    original_monthly = current_principal / remaining_original_months
                    new_months = math.ceil(current_principal / original_monthly)
                else:
                    original_monthly = original_result.records[0].monthly_payment
                    n = math.log(
                        original_monthly / (original_monthly - current_principal * monthly_rate)
                    ) / math.log(1 + monthly_rate)
                    new_months = math.ceil(n)
            else:  # 等额本金
                original_monthly_principal = (
                    original_result.principal / original_result.total_months
                )
                new_months = math.ceil(current_principal / original_monthly_principal)
        else:  # 降低月供
            new_months = remaining_original_months

        # ---- 生成提前还款后的还款计划 ----
        if method == "等额本息":
            new_segment = _calc_equal_principal_and_interest(
                current_principal,
                annual_rate,
                monthly_rate,
                new_months,
                start_period=prepay_period + 1,
                is_prepayment=True,
                prepay_amount=prepay_amount,
                strategy=strategy,
            )
        else:
            new_segment = _calc_equal_principal(
                current_principal,
                annual_rate,
                monthly_rate,
                new_months,
                start_period=prepay_period + 1,
                is_prepayment=True,
                prepay_amount=prepay_amount,
                strategy=strategy,
            )

        full_records.extend(new_segment)

        # ---- 保存详情 ----
        prepay_details.append(
            {
                "次序": idx + 1,
                "提前还款期数": prepay_period,
                "提前还款金额": prepay_amount,
                "策略": strategy,
                "还款前剩余本金": round(remaining_before, 2),
                "新的总期数": prepay_period + new_months,
                "新月供首月": round(new_segment[0].monthly_payment, 2) if new_segment else 0,
            }
        )

        # ---- 更新当前状态 ----
        current_period = prepay_period + 1
        current_records = new_segment
        current_total_months = prepay_period + new_months

    # 重新编号
    for i, r in enumerate(full_records):
        r.period = i + 1

    # 汇总
    final_months = len(full_records)
    final_total_payment = sum(r.monthly_payment for r in full_records) + total_prepay_amount
    final_total_interest = sum(r.interest_paid for r in full_records)
    interest_saved = original_result.total_interest - final_total_interest

    return MultiplePrepaymentResult(
        original_result=original_result,
        prepayments=[
            PrepaymentEvent(p["period"], p["amount"], p["strategy"]) for p in sorted_prepays
        ],
        total_prepay_amount=round(total_prepay_amount, 2),
        final_months=final_months,
        final_total_payment=round(final_total_payment, 2),
        final_total_interest=round(final_total_interest, 2),
        interest_saved=round(interest_saved, 2),
        prepay_details=prepay_details,
        full_records=full_records,
    )


# =============================================================================
# 导出函数
# =============================================================================

def export_to_csv(
    result: LoanResult,
    filepath: str,
    multi_prepay: Optional[MultiplePrepaymentResult] = None,
) -> None:
    """导出还款计划到 CSV 文件。"""
    fieldnames = [
        "期数",
        "月供金额",
        "本月所还本金",
        "本月所还利息",
        "上月剩余本金",
        "上月剩余利息",
        "提前还款金额",
        "策略",
        "备注",
    ]

    records = multi_prepay.full_records if multi_prepay else result.records

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in records:
            note = f"【提前还款-{r.strategy}】" if r.prepay_amount > 0 else ""
            writer.writerow(
                {
                    "期数": r.period,
                    "月供金额": r.monthly_payment,
                    "本月所还本金": r.principal_paid,
                    "本月所还利息": r.interest_paid,
                    "上月剩余本金": r.remaining_principal,
                    "上月剩余利息": r.remaining_interest,
                    "提前还款金额": r.prepay_amount if r.prepay_amount > 0 else "",
                    "策略": r.strategy if r.prepay_amount > 0 else "",
                    "备注": note,
                }
            )

    print(f"✅ CSV 已导出: {os.path.abspath(filepath)}")


def export_to_xlsx(
    result: LoanResult,
    filepath: str,
    multi_prepay: Optional[MultiplePrepaymentResult] = None,
) -> None:
    """导出还款计划到 XLSX 文件。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("⚠️ 未安装 openpyxl，将导出为 CSV 格式。")
        export_to_csv(result, filepath.replace(".xlsx", ".csv"), multi_prepay)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "还款计划"

    headers = [
        "期数",
        "月供金额",
        "本月所还本金",
        "本月所还利息",
        "上月剩余本金",
        "上月剩余利息",
        "提前还款金额",
        "策略",
        "备注",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1783FF", end_color="1783FF", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    prepay_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    after_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    records = multi_prepay.full_records if multi_prepay else result.records
    for r in records:
        note = f"【提前还款-{r.strategy}】" if r.prepay_amount > 0 else ""
        ws.append(
            [
                r.period,
                r.monthly_payment,
                r.principal_paid,
                r.interest_paid,
                r.remaining_principal,
                r.remaining_interest,
                r.prepay_amount if r.prepay_amount > 0 else "",
                r.strategy if r.prepay_amount > 0 else "",
                note,
            ]
        )

        if r.prepay_amount > 0:
            for cell in ws[ws.max_row]:
                cell.fill = prepay_fill
        elif multi_prepay:
            for cell in ws[ws.max_row]:
                cell.fill = after_fill

    # 汇总信息
    ws.append([])
    ws.append(["贷款本金", result.principal])
    ws.append(["合同年利率", f"{result.annual_rate * 100:.4f}%"])
    ws.append(["总期数（月）", result.total_months])
    ws.append(["还款方式", result.method])
    ws.append(["月利率", f"{result.monthly_rate * 100:.4f}%"])
    ws.append(["总还款额", result.total_payment])
    ws.append(["总利息", result.total_interest])

    if multi_prepay:
        ws.append([])
        ws.append(["======== 多次提前还款信息 ========", ""])
        ws.append(["累计提前还款总额", multi_prepay.total_prepay_amount])
        ws.append(["最终总期数", multi_prepay.final_months])
        ws.append(["最终总还款额", multi_prepay.final_total_payment])
        ws.append(["最终总利息", multi_prepay.final_total_interest])
        ws.append(["节省利息", multi_prepay.interest_saved])
        ws.append(
            ["节省比例", f"{multi_prepay.interest_saved / result.total_interest * 100:.2f}%"]
        )

        ws.append([])
        ws.append(["======== 每次提前还款详情 ========", ""])
        for detail in multi_prepay.prepay_details:
            for k, v in detail.items():
                ws.append([k, v])
            ws.append([])

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 25)

    wb.save(filepath)
    print(f"✅ XLSX 已导出: {os.path.abspath(filepath)}")


# =============================================================================
# 打印函数（纯展示，不影响计算）
# =============================================================================

def print_summary(result: LoanResult) -> None:
    """打印贷款计算结果摘要。"""
    print("\n" + "=" * 50)
    print("📊 贷款计算结果")
    print("=" * 50)
    print(f"  贷款本金:     {result.principal:,.2f} 元")
    print(f"  合同年利率:   {result.annual_rate * 100:.4f}%")
    print(f"  月利率:       {result.monthly_rate * 100:.4f}%")
    print(f"  总期数:       {result.total_months} 个月")
    print(f"  还款方式:     {result.method}")
    print("-" * 50)
    print(f"  总还款额:     {result.total_payment:,.2f} 元")
    print(f"  总利息:       {result.total_interest:,.2f} 元")
    print(f"  本利和:       {result.principal + result.total_interest:,.2f} 元")
    print("=" * 50)


def print_schedule(
    result: LoanResult,
    max_rows: int = 12,
    multi_prepay: Optional[MultiplePrepaymentResult] = None,
) -> None:
    """打印还款计划表。"""
    records = multi_prepay.full_records if multi_prepay else result.records
    total = len(records)
    prepay_periods = [p.prepay_period for p in multi_prepay.prepayments] if multi_prepay else []

    print(f"\n📋 还款计划明细（共 {total} 期）")
    print("-" * 110)
    print(
        f"{'期数':>4} {'月供金额':>12} {'还本金':>12} {'还利息':>12} "
        f"{'剩余本金':>14} {'剩余利息':>14} {'提前还款':>12} {'备注':>12}"
    )
    print("-" * 110)

    for r in records[:max_rows]:
        note = "【提前还款】" if r.prepay_amount > 0 else ""
        print(
            f"{r.period:>4} {r.monthly_payment:>12,.2f} {r.principal_paid:>12,.2f} "
            f"{r.interest_paid:>12,.2f} {r.remaining_principal:>14,.2f} "
            f"{r.remaining_interest:>14,.2f} {r.prepay_amount:>12,.2f} {note:>12}"
        )

    if total > max_rows * 2:
        print("...")
        for pp in prepay_periods:
            if max_rows < pp < total - max_rows:
                for r in records[max(0, pp - 2) : min(pp + 5, total)]:
                    note = "【提前还款】" if r.prepay_amount > 0 else ""
                    print(
                        f"{r.period:>4} {r.monthly_payment:>12,.2f} {r.principal_paid:>12,.2f} "
                        f"{r.interest_paid:>12,.2f} {r.remaining_principal:>14,.2f} "
                        f"{r.remaining_interest:>14,.2f} {r.prepay_amount:>12,.2f} {note:>12}"
                    )
                print("...")
                break

    for r in records[-max_rows:]:
        note = "【提前还款】" if r.prepay_amount > 0 else ""
        print(
            f"{r.period:>4} {r.monthly_payment:>12,.2f} {r.principal_paid:>12,.2f} "
            f"{r.interest_paid:>12,.2f} {r.remaining_principal:>14,.2f} "
            f"{r.remaining_interest:>14,.2f} {r.prepay_amount:>12,.2f} {note:>12}"
        )

    print("-" * 110)


def print_multiple_prepayment_summary(multi_prepay: MultiplePrepaymentResult) -> None:
    """打印多次提前还款摘要。"""
    orig = multi_prepay.original_result

    print("\n" + "=" * 60)
    print("💰 多次提前还款分析")
    print("=" * 60)
    print(f"  原始贷款总利息:   {orig.total_interest:,.2f} 元")
    print(f"  累计提前还款:     {multi_prepay.total_prepay_amount:,.2f} 元")
    print(f"  最终总期数:       {multi_prepay.final_months} 个月（原 {orig.total_months} 个月）")
    print(f"  最终总还款额:     {multi_prepay.final_total_payment:,.2f} 元")
    print(f"  最终总利息:       {multi_prepay.final_total_interest:,.2f} 元")
    print("-" * 60)
    print(f"  【节省利息】      {multi_prepay.interest_saved:,.2f} 元")
    print(f"  【节省比例】      {multi_prepay.interest_saved / orig.total_interest * 100:.2f}%")
    print("=" * 60)

    print("\n📋 每次提前还款详情:")
    print("-" * 60)
    for detail in multi_prepay.prepay_details:
        print(f"\n  第 {detail['次序']} 次提前还款:")
        for k, v in detail.items():
            if k != "次序":
                print(f"    {k}: {v}")


# =============================================================================
# 交互式命令行
# =============================================================================

def _input_float(prompt: str, min_val: float = 0) -> float:
    """安全输入浮点数。"""
    while True:
        try:
            val = float(input(prompt).strip().replace(",", ""))
            if val <= min_val:
                print(f"❌ 必须大于 {min_val}")
                continue
            return val
        except ValueError:
            print("❌ 请输入有效的数字")


def _input_int(prompt: str, min_val: int = 0, max_val: Optional[int] = None) -> int:
    """安全输入整数。"""
    while True:
        try:
            val = int(input(prompt).strip())
            if val <= min_val:
                print(f"❌ 必须大于 {min_val}")
                continue
            if max_val is not None and val > max_val:
                print(f"❌ 必须小于等于 {max_val}")
                continue
            return val
        except ValueError:
            print("❌ 请输入有效的整数")


def interactive_mode() -> None:
    """交互式命令行模式。"""
    print("🏦 欢迎使用贷款计算器（含多次提前还款功能）")
    print("-" * 40)

    principal = _input_float("请输入贷款本金（元）: ")

    while True:
        rate_input = input("请输入合同年利率（如 4.3 表示 4.3%）: ").strip().replace("%", "")
        try:
            annual_rate = float(rate_input)
            if annual_rate < 0:
                print("❌ 利率不能为负数")
                continue
            if annual_rate > 1:
                annual_rate = annual_rate / 100
            break
        except ValueError:
            print("❌ 请输入有效的数字")

    total_months = _input_int("请输入总期数（月）: ", min_val=0)

    while True:
        method = input("请选择还款方式 [1-等额本息 / 2-等额本金]: ").strip()
        if method in ("1", "等额本息"):
            method = "等额本息"
            break
        if method in ("2", "等额本金"):
            method = "等额本金"
            break
        print("❌ 请输入 1 或 2")

    # 计算原始贷款
    result = calculate_loan(principal, annual_rate, total_months, method)
    print_summary(result)
    print_schedule(result, max_rows=6)

    # 提前还款
    prepayments: List[Dict] = []
    if input("\n是否进行提前还款计算? [y/n]: ").strip().lower() in ("y", "yes", "是"):
        while True:
            print(f"\n--- 第 {len(prepayments) + 1} 次提前还款 ---")

            min_period = (prepayments[-1]["period"] + 1) if prepayments else 1
            prepay_period = _input_int(
                f"请输入提前还款期数（第几期后，{min_period}-{total_months - 1}）: ",
                min_val=min_period - 1,
                max_val=total_months - 1,
            )

            prepay_amount = _input_float("请输入提前还款金额（元）: ")

            while True:
                strategy = input(
                    "请选择还款策略 [1-缩短期限（月供不变）/ 2-降低月供（期数不变）]: "
                ).strip()
                if strategy in ("1", "缩短期限"):
                    strategy = "缩短期限"
                    break
                if strategy in ("2", "降低月供"):
                    strategy = "降低月供"
                    break
                print("❌ 请输入 1 或 2")

            prepayments.append(
                {"period": prepay_period, "amount": prepay_amount, "strategy": strategy}
            )

            if input("是否继续添加提前还款? [y/n]: ").strip().lower() not in (
                "y",
                "yes",
                "是",
            ):
                break

        multi_prepay = (
            calculate_multiple_prepayments(result, prepayments) if prepayments else None
        )
        if multi_prepay:
            print_multiple_prepayment_summary(multi_prepay)
            print_schedule(result, max_rows=6, multi_prepay=multi_prepay)
    else:
        multi_prepay = None

    # 导出
    while True:
        fmt = input("\n请选择导出格式 [1-CSV / 2-XLSX / 0-不导出]: ").strip()
        if fmt in ("0", "不导出"):
            break
        if fmt in ("1", "csv", "CSV"):
            fmt = "csv"
            break
        if fmt in ("2", "xlsx", "XLSX"):
            fmt = "xlsx"
            break
        print("❌ 请输入 0、1 或 2")

    if fmt != "0":
        default_name = f"还款计划_{method}_{total_months}期"
        if multi_prepay:
            default_name += "_多次提前还款"
        default_name += f".{fmt}"

        filepath = input(f"请输入保存路径（回车使用默认: {default_name}）: ").strip()
        if not filepath:
            filepath = default_name

        if fmt == "csv":
            export_to_csv(result, filepath, multi_prepay)
        else:
            export_to_xlsx(result, filepath, multi_prepay)

    print("\n🎉 计算完成！")


def demo() -> None:
    """演示模式：展示模块 API 用法。"""
    print("🚀 运行演示模式...\n")

    # 1. 计算原始贷款
    result = calculate_loan(
        principal=650000,
        annual_rate=0.0391,
        total_months=324,
        method="等额本息",
    )
    print_summary(result)

    # 2. 定义多次提前还款计划（用户输入的参数）
    prepayments = [
        {"period": 60, "amount": 100000, "strategy": "缩短期限"},
        {"period": 100, "amount": 100000, "strategy": "降低月供"},
    ]

    # 3. 计算多次提前还款后的结果
    multi_prepay = calculate_multiple_prepayments(result, prepayments)
    print_multiple_prepayment_summary(multi_prepay)
    print_schedule(result, max_rows=4, multi_prepay=multi_prepay)

    # 4. 导出
    export_to_csv(result, "demo_多次提前还款.csv", multi_prepay)
    export_to_xlsx(result, "demo_多次提前还款.xlsx", multi_prepay)


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] in ("--demo", "-d", "demo"):
        demo()
    else:
        interactive_mode()
